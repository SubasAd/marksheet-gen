from openpyxl import load_workbook
import re

def read_students(sheet_name, excel_file, start_row=11):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name]
    merged_cell_map = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        value = top_left_cell.value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if value: # Only map if there is a value
                    merged_cell_map[(row, col)] = str(value).strip()

    # --- Now, read the headers using this map ---
    header_rows = range(6, 11) # Your header rows
    headers = []
    for col in range(1, ws.max_column + 1):
        parts = []
        for row in header_rows:
            # Check if this cell is part of a merged range
            if (row, col) in merged_cell_map:
                cell_value = merged_cell_map[(row, col)]
            else:
                cell_value = ws.cell(row=row, column=col).value
            
            if cell_value is not None and str(cell_value).strip():
                part = str(cell_value).strip()
                if part not in parts:
                    parts.append(part)
        
        headers.append(" - ".join(parts))
    students = []
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=True):
        if all(cell is None for cell in row):
            continue
        students.append({header: value for header, value in zip(headers, row)})

    wb.close()
    return students


def to_snake_case(text: str) -> str:
    """Convert arbitrary header text to snake_case."""
    if not text:
        return ""
    text = re.sub(r"[^0-9a-zA-Z]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_").lower()
    return text

def read_school_info(excel_file, sheet_name="School Details"):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name]

    headers = [to_snake_case(cell.value) for cell in ws[1] if cell.value]

    values = [cell.value for cell in ws[2][:len(headers)]]

    wb.close()

    return dict(zip(headers, values))